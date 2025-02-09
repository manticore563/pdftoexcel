import tabula
import pandas as pd
import logging
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def process_pdf_to_excel(pdf_path, excel_path):
    """
    Process a PDF file and convert its content to Excel format while preserving formatting
    """
    try:
        logging.info(f"Starting PDF processing: {pdf_path}")

        # Extract tables from PDF using both modes to ensure we catch all table structures
        tables = []

        # Try lattice mode first
        logging.info("Attempting table extraction with lattice mode")
        lattice_tables = tabula.read_pdf(
            pdf_path,
            pages='all',
            multiple_tables=True,
            lattice=True,
            guess=True,
            silent=True
        )

        if lattice_tables:
            tables.extend(lattice_tables)
            logging.info(f"Found {len(lattice_tables)} tables in lattice mode")

        # Try stream mode if needed
        if not tables or all(table.empty for table in tables):
            logging.info("Attempting table extraction with stream mode")
            stream_tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                stream=True,
                guess=True,
                silent=True
            )
            if stream_tables:
                tables = stream_tables
                logging.info(f"Found {len(stream_tables)} tables in stream mode")

        if not tables:
            raise ValueError("No table data found in PDF")

        # Process each table separately
        processed_tables = []
        max_columns = 0

        for idx, table in enumerate(tables):
            if table.empty:
                logging.info(f"Skipping empty table {idx}")
                continue

            # Clean the table
            table = table.dropna(how='all', axis=0).dropna(how='all', axis=1)

            if table.empty:
                continue

            # Track maximum number of columns
            max_columns = max(max_columns, len(table.columns))

            # Convert all columns to string type to prevent type mismatches
            table = table.astype(str)

            # Store cleaned table
            processed_tables.append(table)
            logging.info(f"Processed table {idx}: {len(table.columns)} columns, {len(table)} rows")

        if not processed_tables:
            raise ValueError("No valid tables found after processing")

        # Standardize all tables to have the same number of columns
        final_tables = []
        for idx, table in enumerate(processed_tables):
            if len(table.columns) < max_columns:
                # Add missing columns with empty strings
                for i in range(len(table.columns), max_columns):
                    table[f'Column_{i+1}'] = ''
            # Ensure consistent column names
            table.columns = [f'Column_{i+1}' for i in range(max_columns)]
            final_tables.append(table)
            logging.info(f"Standardized table {idx} to {max_columns} columns")

        # Combine all tables
        df = pd.concat(final_tables, ignore_index=True)
        df = df.fillna('')  # Clean up any remaining NaN values
        logging.info(f"Combined {len(final_tables)} tables into final dataset with {len(df)} rows")

        # Save to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply formatting
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format headers
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-adjust column width
                column_letter = get_column_letter(col)
                max_length = max(
                    len(str(cell.value)) for cell in worksheet[column_letter]
                    if cell.value is not None
                )
                worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)

            # Format data cells
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row, col)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            logging.info(f"Successfully saved Excel file with formatting: {excel_path}")

    except Exception as e:
        logging.error(f"Error processing PDF: {str(e)}")
        raise Exception(f"Failed to process PDF: {str(e)}")